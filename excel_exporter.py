#!/usr/bin/env python3
"""
Excel Exporter for Consolidated Reports
========================================

Creates multi-sheet Excel workbooks with Summary, Details, Rollup, and
Adjustment Mapping data.

Dependencies: openpyxl
"""

import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _apply_header_formatting(ws, num_cols):
    """Apply header formatting to the first row."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def _apply_data_formatting(ws, num_rows, num_cols):
    """Apply data cell formatting."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row_num in range(2, num_rows + 1):
        for col_num in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            
            # Align numbers to right
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                cell.alignment = Alignment(horizontal="right")


def _auto_adjust_columns(ws):
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value or "")) > max_length:
                    max_length = len(str(cell.value or ""))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_summary_details_excel(output_dir, filename_prefix, summary_data, details_data, 
                                  rollup_data=None, adjustment_mapping=None):
    """
    Create a multi-sheet Excel workbook with Summary, Details, Rollup, and 
    Adjustment_Mapping sheets.
    
    Args:
        output_dir: Directory to write the Excel file
        filename_prefix: Prefix for the filename (will add timestamp and .xlsx)
        summary_data: List of dicts for Summary sheet
        details_data: List of dicts for Details sheet
        rollup_data: List of dicts for Rollup sheet (optional)
        adjustment_mapping: List of dicts for Adjustment_Mapping sheet (optional)
    
    Returns:
        Path to the created Excel file
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    # Create workbook
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # ====================================================================
    # SUMMARY SHEET
    # ====================================================================
    if summary_data and len(summary_data) > 0:
        # Write headers (use union of keys across all rows to capture all columns)
        headers = []
        for row in summary_data:
            for k in row.keys():
                if k not in headers:
                    headers.append(k)
        for col_num, header in enumerate(headers, 1):
            ws_summary.cell(row=1, column=col_num, value=header)
        
        # Write data rows
        for row_num, row_data in enumerate(summary_data, 2):
            for col_num, header in enumerate(headers, 1):
                ws_summary.cell(row=row_num, column=col_num, value=row_data.get(header, ""))
        
        # Apply formatting
        _apply_header_formatting(ws_summary, len(headers))
        _apply_data_formatting(ws_summary, len(summary_data) + 1, len(headers))
        _auto_adjust_columns(ws_summary)
    
    # ====================================================================
    # DETAILS SHEET
    # ====================================================================
    ws_details = wb.create_sheet("Details")
    if details_data and len(details_data) > 0:
        # Write headers (use union of keys across all rows to capture all columns)
        headers = []
        for row in details_data:
            for k in row.keys():
                if k not in headers:
                    headers.append(k)
        for col_num, header in enumerate(headers, 1):
            ws_details.cell(row=1, column=col_num, value=header)
        
        # Write data rows
        for row_num, row_data in enumerate(details_data, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                ws_details.cell(row=row_num, column=col_num, value=value)
        
        # Apply formatting
        _apply_header_formatting(ws_details, len(headers))
        _apply_data_formatting(ws_details, len(details_data) + 1, len(headers))
        _auto_adjust_columns(ws_details)
    
    # ====================================================================
    # ROLLUP SHEET (if provided)
    # ====================================================================
    if rollup_data:
        ws_rollup = wb.create_sheet("Rollup")
        if len(rollup_data) > 0:
            # Write headers (use union of keys across all rows to capture all columns)
            headers = []
            for row in rollup_data:
                for k in row.keys():
                    if k not in headers:
                        headers.append(k)

            for col_num, header in enumerate(headers, 1):
                ws_rollup.cell(row=1, column=col_num, value=header)

            # Write data rows
            for row_num, row_data in enumerate(rollup_data, 2):
                for col_num, header in enumerate(headers, 1):
                    value = row_data.get(header, "")
                    ws_rollup.cell(row=row_num, column=col_num, value=value)

            # Apply formatting
            _apply_header_formatting(ws_rollup, len(headers))
            _apply_data_formatting(ws_rollup, len(rollup_data) + 1, len(headers))
            _auto_adjust_columns(ws_rollup)
    
    # ====================================================================
    # ADJUSTMENT MAPPING SHEET (if provided)
    # ====================================================================
    if adjustment_mapping:
        ws_adjustment = wb.create_sheet("Adjustment_Mapping")
        if len(adjustment_mapping) > 0:
            # Write headers (use union of keys across all rows to capture all columns)
            headers = []
            for row in adjustment_mapping:
                for k in row.keys():
                    if k not in headers:
                        headers.append(k)
            for col_num, header in enumerate(headers, 1):
                ws_adjustment.cell(row=1, column=col_num, value=header)
            
            # Write data rows
            for row_num, row_data in enumerate(adjustment_mapping, 2):
                for col_num, header in enumerate(headers, 1):
                    ws_adjustment.cell(row=row_num, column=col_num, value=row_data.get(header, ""))
            
            # Apply formatting
            _apply_header_formatting(ws_adjustment, len(headers))
            _apply_data_formatting(ws_adjustment, len(adjustment_mapping) + 1, len(headers))
            _auto_adjust_columns(ws_adjustment)
    
    # ====================================================================
    # SAVE WORKBOOK
    # ====================================================================
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    
    return filepath


if __name__ == "__main__":
    # Test example
    test_summary = [
        {"Practice": "Cardiology", "Patients": 42, "Active": 40, "PreAuth": 8},
        {"Practice": "Orthopedics", "Patients": 35, "Active": 32, "PreAuth": 5},
    ]
    
    test_details = [
        {
            "Patient": "Smith, John",
            "Member_ID": "M001",
            "Copay": 25,
            "Deductible": 500,
            "Coverage_Status": "Active"
        },
        {
            "Patient": "Doe, Jane",
            "Member_ID": "M002",
            "Copay": 35,
            "Deductible": 1000,
            "Coverage_Status": "Active"
        },
    ]
    
    test_rollup = [
        {"Category": "Coverage Status", "Value": "Active", "Count": 72, "Percentage": "88.9%"},
        {"Category": "Network Status", "Value": "In-Network", "Count": 75, "Percentage": "92.6%"},
        {"Category": "Alert Type", "Value": "HIGH_DEDUCTIBLE", "Count": 12, "Percentage": "15.8%"},
    ]
    
    output_path = create_summary_details_excel(
        "test_output",
        "test_report",
        test_summary,
        test_details,
        test_rollup
    )
    print(f"Created: {output_path}")
